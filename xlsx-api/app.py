"""
Instagram Influencer XLSX Generator API
POST /generate — builds influencer profile + audit workbooks from n8n payload.
"""

from __future__ import annotations

import base64
import io
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from pydantic import BaseModel, Field

# Prefer xlsx-api/templates (Render deploy); fallback to repo-root templates/
_API_DIR = Path(__file__).resolve().parent
TEMPLATES_DIR = _API_DIR / "templates"
if not (TEMPLATES_DIR / "influencer_template.xlsx").exists():
    TEMPLATES_DIR = _API_DIR.parent / "templates"
INFLUENCER_TEMPLATE = TEMPLATES_DIR / "influencer_template.xlsx"
AUDIT_TEMPLATE = TEMPLATES_DIR / "audit_template.xlsx"

INFLUENCER_HEADERS = [
    "Full Name",
    "Channel Name",
    "Influencer Category",
    "Gender",
    "Email Id",
    "CONTACT No.",
    "Instagram Link",
    "IG Followers",
    "You Tube Link",
    "YT Subs",
    "Facebook Link",
    "FB Followers",
    "Linkedin Link",
    "Linkedin Followers",
    "Snapchat Link",
    "SC Followers",
    "External Link",
    "Country",
    "State",
    "City",
    "Region",
    "Regional/Primary Language",
    "Influencer Type",
    "Primary Platform",
    "AVG Views (Last 10 Posts)",
]


class GenerateRequest(BaseModel):
    influencer_profile_row: dict[str, Any] = Field(default_factory=dict)
    audit: dict[str, Any] = Field(default_factory=dict)


app = FastAPI(title="Insta XLSX Generator", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def _fmt_num(n: Any) -> str:
    if n is None or n == "":
        return ""
    try:
        return f"{int(float(n)):,}"
    except (TypeError, ValueError):
        return str(n)


def _fmt_pct(rate: float) -> str:
    return f"{rate * 100:.2f}%"


def _username(audit: dict) -> str:
    meta = audit.get("meta") or {}
    return str(meta.get("username") or "profile")


def _workbook_to_base64(wb) -> str:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("ascii")


def build_influencer_workbook(row: dict[str, Any]):
    if not INFLUENCER_TEMPLATE.exists():
        raise FileNotFoundError(f"Missing template: {INFLUENCER_TEMPLATE}")

    wb = load_workbook(INFLUENCER_TEMPLATE)
    ws = wb.active
    for col_idx, header in enumerate(INFLUENCER_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
        val = row.get(header, "")
        if header == "IG Followers" and val != "":
            try:
                val = int(float(val))
            except (TypeError, ValueError):
                pass
        if header == "AVG Views (Last 10 Posts)" and val != "":
            try:
                val = int(float(val))
            except (TypeError, ValueError):
                pass
        ws.cell(row=2, column=col_idx, value=val if val != "" else None)
    return wb


def build_audit_workbook(audit: dict[str, Any]):
    if not AUDIT_TEMPLATE.exists():
        raise FileNotFoundError(f"Missing template: {AUDIT_TEMPLATE}")

    wb = load_workbook(AUDIT_TEMPLATE)
    meta = audit.get("meta") or {}
    metrics = audit.get("metrics") or {}
    posts = audit.get("posts") or []
    insights = audit.get("insights") or {}
    transcripts = audit.get("transcripts") or []
    username = _username(audit)
    window_days = meta.get("window_days", 30)
    generated = meta.get("generated_date") or datetime.utcnow().strftime("%Y-%m-%d")
    total_views = metrics.get("total_views", 0)
    post_count = metrics.get("post_count", len(posts))
    avg_views = metrics.get("avg_views", 0)
    avg_likes = metrics.get("avg_likes", 0)
    avg_er = metrics.get("avg_engagement_rate", 0)

    # --- Dashboard ---
    if "Dashboard" in wb.sheetnames:
        dash = wb["Dashboard"]
        dash["A2"] = f"  @{username.upper()}  -  INSTAGRAM POSTS AUDIT"
        dash["A3"] = (
            f"  Last {window_days} Days  |  {post_count} Posts  |  "
            f"{_fmt_num(total_views)} Total Views  |  Generated {generated}"
        )
        dash["B9"] = post_count
        dash["D9"] = _fmt_num(total_views)
        dash["F9"] = _fmt_num(avg_views)
        dash["H9"] = _fmt_num(avg_likes)
        dash["J9"] = _fmt_pct(float(avg_er)) if avg_er else "0.00%"

        top_posts = audit.get("top_posts") or sorted(
            posts, key=lambda p: p.get("views") or 0, reverse=True
        )[:5]
        header_row = 14
        for i, p in enumerate(top_posts):
            r = header_row + 1 + i
            dash.cell(row=r, column=2, value=i + 1)
            dash.cell(row=r, column=3, value=p.get("url", ""))
            dash.cell(row=r, column=4, value=p.get("date", ""))
            dash.cell(row=r, column=5, value=p.get("views", 0))
            dash.cell(row=r, column=6, value=p.get("likes", 0))
            dash.cell(row=r, column=7, value=p.get("comments", 0))
            er = p.get("engagement_rate_display") or p.get("engagement_rate_pct")
            if er is None and p.get("views"):
                er = f"{(float(p.get('engagement', 0)) / float(p['views']) * 100):.2f}%"
            dash.cell(row=r, column=8, value=er or "0.00%")
            dash.cell(row=r, column=9, value=p.get("status", ""))
            dash.cell(row=r, column=10, value=p.get("topic", ""))
            dash.cell(row=r, column=11, value=p.get("hook_type", ""))

    # --- All Reels (all post types; captions for photos/carousels) ---
    if "All Reels" in wb.sheetnames:
        sheet = wb["All Reels"]
        sheet["A1"] = (
            f"ALL POSTS  ·  @{username}  ·  Last {window_days} Days "
            f"({post_count} posts  ·  {_fmt_num(total_views)} total views)"
        )
        start_row = 5
        for idx, p in enumerate(posts):
            r = start_row + idx
            sheet.cell(row=r, column=2, value=idx + 1)
            sheet.cell(row=r, column=3, value=p.get("date", ""))
            sheet.cell(row=r, column=4, value=p.get("type", "Post"))
            sheet.cell(row=r, column=5, value=p.get("url", ""))
            sheet.cell(row=r, column=6, value=p.get("views", 0))
            sheet.cell(row=r, column=7, value=p.get("likes", 0))
            sheet.cell(row=r, column=8, value=p.get("comments", 0))
            sheet.cell(row=r, column=9, value=p.get("engagement", 0))
            er_pct = p.get("engagement_rate_pct")
            sheet.cell(row=r, column=10, value=er_pct if er_pct is not None else 0)
            sheet.cell(row=r, column=11, value=p.get("status", ""))
            sheet.cell(row=r, column=12, value=p.get("topic", ""))
            sheet.cell(row=r, column=13, value=p.get("hook_type", ""))
            sheet.cell(row=r, column=14, value=p.get("caption_preview", ""))

    # --- Insights ---
    if "Insights" in wb.sheetnames:
        ins = wb["Insights"]
        row = 4
        sections = insights.get("sections") or []
        for section_title, rows in sections:
            ins.cell(row=row, column=2, value=section_title)
            row += 2
            for line in rows:
                if not isinstance(line, list):
                    continue
                for col_offset, cell_val in enumerate(line):
                    if cell_val == "":
                        continue
                    ins.cell(row=row, column=2 + col_offset * 2, value=cell_val)
                row += 1
            row += 1

    # --- Transcripts (reels only) ---
    if "Transcripts" in wb.sheetnames:
        tr = wb["Transcripts"]
        reel_count = len(transcripts)
        tr["A1"] = (
            f"REEL TRANSCRIPTS  ·  @{username}  ·  Last {window_days} Days  ·  "
            f"{reel_count} Reels"
        )
        summary = audit.get("transcript_summary", "")
        if summary:
            tr["B4"] = summary
        start_row = 6
        for idx, t in enumerate(transcripts):
            r = start_row + idx
            tr.cell(row=r, column=2, value=idx + 1)
            tr.cell(row=r, column=3, value=t.get("date", ""))
            tr.cell(row=r, column=4, value=t.get("url", ""))
            tr.cell(row=r, column=5, value=t.get("views", 0))
            tr.cell(row=r, column=6, value=t.get("source", "Caption"))
            text = t.get("text", "")
            tr.cell(row=r, column=7, value=text[:32000] if text else "")

    return wb


@app.get("/")
def root():
    """Render health check + wake endpoint."""
    return {
        "service": "insta-xlsx-api",
        "status": "ok",
        "endpoints": {"health": "/health", "generate": "POST /generate"},
        "templates": {
            "influencer": INFLUENCER_TEMPLATE.exists(),
            "audit": AUDIT_TEMPLATE.exists(),
        },
    }


@app.get("/health")
def health():
    return {
        "ok": True,
        "templates": {
            "influencer": INFLUENCER_TEMPLATE.exists(),
            "audit": AUDIT_TEMPLATE.exists(),
        },
    }


@app.post("/generate")
def generate(req: GenerateRequest):
    if not req.audit:
        raise HTTPException(status_code=400, detail="Missing audit payload")

    try:
        inf_wb = build_influencer_workbook(req.influencer_profile_row)
        aud_wb = build_audit_workbook(req.audit)
    except FileNotFoundError as e:
        raise HTTPException(status_code=500, detail=str(e)) from e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Workbook build failed: {e}") from e

    username = re.sub(r"[^\w.-]", "_", _username(req.audit))
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
    inf_name = f"influencer_{username}_{ts}.xlsx"
    aud_name = f"insta_audit_{username}_{ts}.xlsx"

    return {
        "ok": True,
        "influencer_filename": inf_name,
        "audit_filename": aud_name,
        "influencer_xlsx_base64": _workbook_to_base64(inf_wb),
        "audit_xlsx_base64": _workbook_to_base64(aud_wb),
    }
